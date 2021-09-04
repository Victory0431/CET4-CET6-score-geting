
from lxml import etree
import requests
from openpyxl import load_workbook
from fake_useragent import UserAgent
import os
import datetime
import random
import time
import easygui
#import sys
import openpyxl


path1 = os.getcwd()
ua = UserAgent()

def dealwiths(strr):
    tu = ''
    for i in strr:
        if i == '\n' or i == '\t' or i == ' ':
            pass
        else:
            tu += i
    return tu
def cet4(name,sfz):
    s=requests.session()
    headers={
    'Host': 'cache.neea.edu.cn',
    'Connection': 'keep-alive',
    'Cache-Control': 'max-age=0',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': ua.random,
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Referer': 'http://cjcx.neea.edu.cn/',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Origin': 'http://cjcx.neea.edu.cn',
    'Cookie': 'Hm_lvt_dc1d69ab90346d48ee02f18510292577=1629948390; Hm_lpvt_dc1d69ab90346d48ee02f18510292577=1629948694'#JSESSIONID=385E9163AF2634FFD3713859C7259DF6.tomcat1'#0579209D0D30E5843D6FE34C9D01D187.tomcat1'#'JSESSIONID=3C84F48ACFD5DFBE4BE8A05A6A725CA5.tomcat2'
    }
    url4 = 'http://cache.neea.edu.cn/api/latest/results/cet?e=CET_202106_DANGCI&km=1&xm=' + name + '&no=' + sfz + '&v='
    url6 = 'http://cache.neea.edu.cn/api/latest/results/cet?e=CET_202106_DANGCI&km=2&xm=' + name + '&no=' + sfz + '&v=' 

    rs=s.get(url4,headers=headers,verify=False)
    ui = rs.text
    return ui

def cet6(name,sfz):
    s=requests.session()
    headers={
    'Host': 'cache.neea.edu.cn',
    'Connection': 'keep-alive',
    'Cache-Control': 'max-age=0',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': ua.random,
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Referer': 'http://cjcx.neea.edu.cn/',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Origin': 'http://cjcx.neea.edu.cn',
    'Cookie': 'Hm_lvt_dc1d69ab90346d48ee02f18510292577=1629948390; Hm_lpvt_dc1d69ab90346d48ee02f18510292577=1629948694'#JSESSIONID=385E9163AF2634FFD3713859C7259DF6.tomcat1'#0579209D0D30E5843D6FE34C9D01D187.tomcat1'#'JSESSIONID=3C84F48ACFD5DFBE4BE8A05A6A725CA5.tomcat2'
    }
    url4 = 'http://cache.neea.edu.cn/api/latest/results/cet?e=CET_202106_DANGCI&km=1&xm=' + name + '&no=' + sfz + '&v='
    url6 = 'http://cache.neea.edu.cn/api/latest/results/cet?e=CET_202106_DANGCI&km=2&xm=' + name + '&no=' + sfz + '&v=' 

    rs=s.get(url6,headers=headers,verify=False)
    ui = rs.text
    return ui


if os.path.exists('成绩模板.xlsx'):
    pass
else:
    wb = openpyxl.Workbook()
    wb.save('成绩模板.xlsx')
    wb6 = load_workbook('成绩模板.xlsx')
    ws6 = wb6.active
    ws6['A' + str(1)] = '姓名'
    ws6['B' + str(1)] = '等级'
    ws6['C' + str(1)] = '分数'
    ws6['D' + str(1)] = '听力'
    ws6['E' + str(1)] = '阅读'
    ws6['F' + str(1)] = '写作'
    ws6['I' + str(1)] = '身份证号'
    ws6['G' + str(1)] = '准考证号'
    ws6['H' + str(1)] = '成绩单编号'
    wb6.save('成绩模板.xlsx')

    
msg = '请选择即将查询的文件'
filename = easygui.fileopenbox(msg)

wb1 = load_workbook(filename)
ws1 = wb1.active
ender = ws1.max_row+1
number = 2

save = easygui.enterbox('请输入保存文件名') + '.xlsx'

wb2 = load_workbook('成绩模板.xlsx')
ws2 = wb2.active
s=requests.session()
t1 = datetime.datetime.now()

for cs in range(1,ender):
    name = ws1['A' + str(cs)].value
    sfz = str(ws1['B' + str(cs)].value)
    ui4 = eval(cet4(name,sfz))
    ui6 = eval(cet6(name,sfz))
    if ui4['code'] == 200:
        school= ui4["xx"]
        zkzh= ui4["zkzh"]
        sfz= ui4["sfz"]
        score= ui4["score"]
        sco_lc= ui4["sco_lc"]
        sco_rd= ui4["sco_rd"]
        sco_wt= ui4["sco_wt"]
        cjdbh= ui4["id"]

        ws2['A' + str(number)] = name
        ws2['B' + str(number)] = 'CET4'
        ws2['C' + str(number)] = score
        ws2['D' + str(number)] = sco_lc
        ws2['E' + str(number)] = sco_rd
        ws2['F' + str(number)] = sco_wt
        ws2['I' + str(number)] = sfz
        ws2['G' + str(number)] = zkzh
        ws2['H' + str(number)] = cjdbh
        number += 1
        print(name + ' CET4 ' + score)
        
    if ui6['code'] == 200:
        school= ui6["xx"]
        zkzh= ui6["zkzh"]
        sfz= ui6["sfz"]
        score= ui6["score"]
        sco_lc= ui6["sco_lc"]
        sco_rd= ui6["sco_rd"]
        sco_wt= ui6["sco_wt"]
        cjdbh= ui6["id"]

        ws2['A' + str(number)] = name
        ws2['B' + str(number)] = 'CET6'
        ws2['C' + str(number)] = score
        ws2['D' + str(number)] = sco_lc
        ws2['E' + str(number)] = sco_rd
        ws2['F' + str(number)] = sco_wt
        ws2['I' + str(number)] = sfz
        ws2['G' + str(number)] = zkzh
        ws2['H' + str(number)] = cjdbh
        number += 1
        print(name + ' CET6 ' + score)
    if cs%200 == 0:
        print(cs)
        time.sleep(5)
    
wb2.save(save)  
    
        
        
